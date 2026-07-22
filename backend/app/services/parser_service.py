"""Document parser service for PDF, DOCX, and XLSX files."""

import fitz


async def parse_pdf(file_path: str) -> list[dict]:
    """Extract text from a PDF file.

    Returns a list of dicts, one per page, each containing ``page_num``
    (int) and ``text`` (str). Image-only or empty pages yield ``""``.
    """
    doc = fitz.open(file_path)
    pages: list[dict] = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text().strip()
        pages.append({"page_num": page_num + 1, "text": text})
    doc.close()
    return pages


async def parse_docx(file_path: str) -> list[dict]:
    """Extract text from a DOCX file.

    DOCX has no native page concept, so the result is a single page
    with ``page_num=1`` containing all paragraphs joined by double newline.
    """
    from docx import Document as DocxDocument

    doc = DocxDocument(file_path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    full_text = "\n\n".join(paragraphs)
    return [{"page_num": 1, "text": full_text}]


async def parse_xlsx(file_path: str) -> list[dict]:
    """Extract text from an XLSX file.

    Each sheet becomes one "page" with page_num set to the sheet name.
    Rows within a sheet are joined with `` | `` separator.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    pages: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text: list[str] = []
        for row in ws.iter_rows(values_only=True):
            values = [str(v) if v is not None else "" for v in row]
            rows_text.append(" | ".join(values))
        text = "\n".join(rows_text).strip()
        pages.append({"page_num": sheet_name, "text": text})
    wb.close()
    return pages
